
from datetime import datetime
import openpyxl
from textwrap import fill
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Color, colors, Side, GradientFill, Alignment, Border, PatternFill

class WorksheetMaycao:
    def __init__(self, args):
        self.nomeAluno    = args["nomeAluno"]
        self.workbookName = args["workbookName"]
        self.objetivo     = args["objetivo"]
        self.diasDaSemana = args["diasDaSemana"]
        self.observacao   = args["observacao"]
        self.dataDeInicio = args["dataDeInicio"]
        self.treinos      = args['treinos']
        self.tipoTreino   = args['tipoTreino']
        self.posTreino    = args['posTreino']
        self.workbook     = 0
        self.worksheet    = 0
        self.lastLineOfEexercicio = 0

    def createSheet(self):
        self.workbook  = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = self.workbookName   

    def setDimensions(self):
        self.worksheet.column_dimensions['A'].width = 5 
        self.worksheet.column_dimensions['B'].width = 35 
        self.worksheet.column_dimensions['C'].width = 15 
        self.worksheet.column_dimensions['D'].width = 25 
        self.worksheet.column_dimensions['E'].width = 10 
        self.worksheet.column_dimensions['F'].width = 7 
        self.worksheet.column_dimensions['G'].width = 30 
        self.worksheet.column_dimensions['A'].height = 25
        self.worksheet.column_dimensions['B'].height = 25
        self.worksheet.column_dimensions['C'].height = 25
        self.worksheet.column_dimensions['D'].height = 25
        self.worksheet.column_dimensions['E'].height = 25
        self.worksheet.column_dimensions['F'].height = 25
        self.worksheet.column_dimensions['G'].height = 25

    def appendDadosAluno(self): 
        self.worksheet['C1'] = "Aluno:"
        self.worksheet['C2'] = "Data de Ínicio:"
        self.worksheet['C3'] = "Objetivo:"
        self.worksheet['C4'] = "Dias da semana:"
        self.worksheet['C5'] = "Observação:"
        self.worksheet['C6'] = "Treino:"
        self.worksheet['D1'] = self.nomeAluno
        self.worksheet['D2'] = self.dataDeInicio
        self.worksheet['D3'] = self.objetivo
        self.worksheet['D4'] = self.diasDaSemana
        self.worksheet['D5'] = self.observacao
        self.worksheet['D6'] = self.tipoTreino
    
    def stylesDadosAluno(self):
        key_font = Font(color='00FFFFFF')
        key_fill = PatternFill("solid", fgColor="00969696")
        key_alignment = Alignment(horizontal="right", vertical="center")     
        self.worksheet['C1'].border = Border(top = Side(border_style="thick", color="00000000"),
        left = Side(border_style="thick", color="00000000"))
        self.worksheet['C2'].border = Border(left = Side(border_style="thick", color="00000000"))
        self.worksheet['C3'].border = Border(left = Side(border_style="thick", color="00000000"))
        self.worksheet['C4'].border = Border(left = Side(border_style="thick", color="00000000"))
        self.worksheet['C5'].border = Border(left = Side(border_style="thick", color="00000000"))

        

        self.worksheet['D2'].border = Border(right = Side(border_style="thick", color="00000000"))
        self.worksheet['D3'].border = Border(right = Side(border_style="thick", color="00000000"))
        self.worksheet['D4'].border = Border(right = Side(border_style="thick", color="00000000"))
        self.worksheet['D5'].border = Border(right = Side(border_style="thick", color="00000000"))

        self.worksheet['C6'].border = Border(left = Side(border_style="thick", color="00000000"),
                                            bottom = Side(border_style="thick", color="00000000"))
        self.worksheet['D1'].border = Border(top = Side(border_style="thick", color="00000000"),
                                            right = Side(border_style="thick", color="00000000"))
        self.worksheet['D6'].border = Border(right = Side(border_style="thick", color="00000000"),
                                                        bottom = Side(border_style="thick", color="00000000"))

        count = 1
        while count < 7:
            cel = 'C' + str(count)
            print(cel)
            self.worksheet[cel].font = key_font
            self.worksheet[cel].fill = key_fill
            self.worksheet[cel].alignment = key_alignment
            count = count + 1

        value_font = Font(color='00000000', bold=True)
        value_fill = PatternFill("solid", fgColor="00FFFFFF")
        value_alignment = Alignment(horizontal="center", vertical="center")

        count = 1
        while count < 7:
            cel = 'D' + str(count)
            self.worksheet[cel].font = value_font
            self.worksheet[cel].fill = value_fill
            self.worksheet[cel].alignment = value_alignment
            count = count + 1

    def appendExercicios(self):
        startingCelLine = 7
        
        for i, treino in enumerate(self.treinos):
            startingCelLine+= 1
            treinoStartingCelLineCopy = startingCelLine + 1
            columnNumber = str(startingCelLine)
            self.worksheet['A' + columnNumber] = "Nº"
            self.worksheet['B' + columnNumber] = "EXERCÍCIO"
            self.worksheet['C' + columnNumber] = "SÉRIES"
            self.worksheet['D' + columnNumber] = "REPETIÇÕES"
            self.worksheet['E' + columnNumber] = "DESCANSO"
            self.worksheet['F' + columnNumber] = "VÍDEO"
            self.worksheet['G' + columnNumber] = "OBS"

            self.styleExercicioColumns(columnNumber)
            
            for iExercicio, exercicio in enumerate(treino.exercicios):
                startingCelLine+= 1
                columnNumber = str(startingCelLine)
                self.worksheet['A' + columnNumber] = iExercicio + 1
                self.worksheet['B' + columnNumber] = exercicio.nomeExercicio
                self.worksheet['C' + columnNumber] = exercicio.series
                self.worksheet['D' + columnNumber] = exercicio.repeticoes
                self.worksheet['E' + columnNumber] = exercicio.descanso
                self.worksheet['F' + columnNumber].hyperlink = exercicio.linkVideo
                self.worksheet['F' + columnNumber].value = ''
                img = openpyxl.drawing.image.Image('./ytb.png')
                img.width = 20
                img.height = 20
                img.anchor = 'F' + columnNumber
                self.worksheet.add_image(img)
                self.worksheet['G' + columnNumber] = exercicio.observacao
                self.styleExercicioRows(columnNumber)
                treinoEndingCelLineCopy = startingCelLine -1


            self.worksheet.merge_cells('H' + str(treinoStartingCelLineCopy) + ':H' + str(columnNumber))
            self.worksheet['H' + str(treinoStartingCelLineCopy)].alignment = Alignment(horizontal="center", vertical="center")
            self.worksheet['H' + str(treinoStartingCelLineCopy)].fill = PatternFill("solid", fgColor="00333333")
            print('\tAdicionei um exercicio \n \n')
            self.worksheet['H' + str(treinoStartingCelLineCopy)].font = Font(color='00FFFF00',size=20)
            self.worksheet['H' + str(treinoStartingCelLineCopy)] = treino.dia  
        self.lastLineOfEexercicio = int(columnNumber)     

    def styleExercicioRows(self, line) : 
        alignment = Alignment(horizontal="center", vertical="center")
        coluns = ['A','B','C','D','E','F','G', 'H']
        for column in coluns : 
            self.worksheet[column + line].alignment = alignment
        self.worksheet['H' + line].font = Font(color='00FFFF00',size=20)

    def styleExercicioColumns(self, line) :
        fill = PatternFill("solid", fgColor="00333333")
        font = Font(color='00FFFF00')
        alignment = Alignment(horizontal="center", vertical="center")
        coluns = ['A','B','C','D','E','F','G', 'H']
        
        for column in coluns:
            if(column != 'H'):
                self.worksheet[column + line].fill = fill
            self.worksheet[column + line].font = font
            self.worksheet[column + line].alignment = alignment
        
          
    def appendPosTreino(self):
        startingLine = self.lastLineOfEexercicio + 3;
        count = startingLine;
        endLine = startingLine + len(self.posTreino)
        titleCel = 'A' + str( startingLine - 1 )  
        self.worksheet[titleCel] = 'PÓS TREINO TODOS OS DIAS'
        
        while count != endLine :
            cellCount = 'A' + str(count)
            cellNome  = 'B' + str(count) 
            cellLink  = 'C' + str(count) 
            exercicio = self.posTreino[count - 45]
            self.worksheet[cellCount] = count - 44
            self.worksheet[cellNome] = exercicio['nome']
            self.worksheet[cellLink].hyperlink = exercicio['link']
            self.worksheet[cellLink].value = ''
            img = openpyxl.drawing.image.Image('./ytb.png')
            img.width = 20
            img.height = 20
            img.anchor = cellLink
            self.worksheet.add_image(img)
            count = count + 1
        print('\tAdicionei o pos trieno')

        return 1;

    def stylePosTreino(self):
        startingLine = self.lastLineOfEexercicio + 2;
        count = startingLine
        endLine = startingLine + 1 + len(self.posTreino)
        fill = PatternFill("solid", fgColor="00333333")
        font = Font(color='00FFFF00')
        alignment = Alignment(horizontal="center", vertical="center")
        coluns = ['A','B','C']
        for colunm in coluns : 
            self.worksheet[colunm + str(startingLine)].fill = fill
            self.worksheet[colunm + str(startingLine)].font = font
            self.worksheet[colunm + str(startingLine)].alignment = alignment

        while count != endLine:
            celA = "A" + str(count)
            celB = 'B' + str(count)
            celC = 'D' + str(count)
            self.worksheet[celA].alignment = alignment
            self.worksheet[celB].alignment = alignment
            self.worksheet[celC].alignment = alignment
            count = count + 1
        self.worksheet.merge_cells('A' + str(startingLine) + ":C" + str(startingLine))
       

class Exercicio: 
    def __init__(self, args):
        self.nomeExercicio = args["nomeExercicio"]
        self.series        = args["series"]
        self.repeticoes    = args["repeticoes"]
        self.descanso      = args["descanso"]
        self.linkVideo     = args["linkVideo"]
        self.observacao    = args["observacao"]

class Treino:
    def __init__(self, args):
        self.exercicios = args['exercicios'] 
        self.dia = args['dia']

        

def main():
    exerciciosSegunda = {
        "dia" : "SEG",
        "exercicios" :[
            Exercicio({
                "nomeExercicio" : "VOADOR",
                "series" : "4",
                "repeticoes" : "8 A 12",
                "descanso" : "1 MIN",
                "linkVideo" : "https://www.youtube.com/watch?v=p0Mql-QHnvY",
                "observacao" : "DROP SET NA ULTIMA SÉRIE"
            }),
            Exercicio({
                "nomeExercicio" : "ELEVAÇÃO LATERAL",
                "series" : "4",
                "repeticoes" : "8 A 12",
                "descanso" : "1 MIN",
                "linkVideo" : "https://www.youtube.com/watch?v=p0Mql-QHnvY",
                "observacao" : "DROP SET NA ULTIMA SÉRIE"
            }),
            Exercicio({
                "nomeExercicio" : "LEG MAQUINA",
                "series" : "3",
                "repeticoes" : "8 A 12",
                "descanso" : "1 MIN",
                "linkVideo" : "https://www.youtube.com/watch?v=p0Mql-QHnvY",
                "observacao" : "DROP SET NA ULTIMA SÉRIE"
            }),
            Exercicio({
                "nomeExercicio" : "EXTENSORA",
                "series" : "3",
                "repeticoes" : "8 A 12",
                "descanso" : "1 MIN",
                "linkVideo" : "https://www.youtube.com/watch?v=p0Mql-QHnvY",
                "observacao" : "DROP SET NA ULTIMA SÉRIE"
            }),
            Exercicio({
                "nomeExercicio" : "GLUTEO POLIA BAIXA",
                "series" : "4",
                "repeticoes" : "8 A 12",
                "descanso" : "1 MIN",
                "linkVideo" : "https://www.youtube.com/watch?v=p0Mql-QHnvY",
                "observacao" : "DROP SET NA ULTIMA SÉRIE"
            }),
            Exercicio({
                "nomeExercicio" : "PANTURRILHA SENTADA",
                "series" : "4",
                "repeticoes" : "8 A 12",
                "descanso" : "1 MIN",
                "linkVideo" : "https://www.youtube.com/watch?v=p0Mql-QHnvY",
                "observacao" : "DROP SET NA ULTIMA SÉRIE"
            }),
        ]
    }

    exerciciosTerca = {
        "dia" : "TER",
        "exercicios" :[
            Exercicio({
                "nomeExercicio" : "VOADOR",
                "series" : "4",
                "repeticoes" : "8 A 12",
                "descanso" : "1 MIN",
                "linkVideo" : "https://www.youtube.com/watch?v=p0Mql-QHnvY",
                "observacao" : "DROP SET NA ULTIMA SÉRIE"
            }),
            Exercicio({
                "nomeExercicio" : "ELEVAÇÃO LATERAL",
                "series" : "4",
                "repeticoes" : "8 A 12",
                "descanso" : "1 MIN",
                "linkVideo" : "https://www.youtube.com/watch?v=p0Mql-QHnvY",
                "observacao" : "DROP SET NA ULTIMA SÉRIE"
            }),
            Exercicio({
                "nomeExercicio" : "LEG MAQUINA",
                "series" : "3",
                "repeticoes" : "8 A 12",
                "descanso" : "1 MIN",
                "linkVideo" : "https://www.youtube.com/watch?v=p0Mql-QHnvY",
                "observacao" : "DROP SET NA ULTIMA SÉRIE"
            }),
            Exercicio({
                "nomeExercicio" : "EXTENSORA",
                "series" : "3",
                "repeticoes" : "8 A 12",
                "descanso" : "1 MIN",
                "linkVideo" : "https://www.youtube.com/watch?v=p0Mql-QHnvY",
                "observacao" : "DROP SET NA ULTIMA SÉRIE"
            }),
            Exercicio({
                "nomeExercicio" : "GLUTEO POLIA BAIXA",
                "series" : "4",
                "repeticoes" : "8 A 12",
                "descanso" : "1 MIN",
                "linkVideo" : "https://www.youtube.com/watch?v=p0Mql-QHnvY",
                "observacao" : "DROP SET NA ULTIMA SÉRIE"
            }),
            Exercicio({
                "nomeExercicio" : "PANTURRILHA SENTADA",
                "series" : "4",
                "repeticoes" : "8 A 12",
                "descanso" : "1 MIN",
                "linkVideo" : "https://www.youtube.com/watch?v=p0Mql-QHnvY",
                "observacao" : "DROP SET NA ULTIMA SÉRIE"
            }),
        ]
    }

    exerciciosQuarta = {
        "dia" : "QUA",
        "exercicios" :[
            Exercicio({
                "nomeExercicio" : "VOADOR",
                "series" : "4",
                "repeticoes" : "8 A 12",
                "descanso" : "1 MIN",
                "linkVideo" : "https://www.youtube.com/watch?v=p0Mql-QHnvY",
                "observacao" : "DROP SET NA ULTIMA SÉRIE"
            }),
            Exercicio({
                "nomeExercicio" : "ELEVAÇÃO LATERAL",
                "series" : "4",
                "repeticoes" : "8 A 12",
                "descanso" : "1 MIN",
                "linkVideo" : "https://www.youtube.com/watch?v=p0Mql-QHnvY",
                "observacao" : "DROP SET NA ULTIMA SÉRIE"
            }),
            Exercicio({
                "nomeExercicio" : "LEG MAQUINA",
                "series" : "3",
                "repeticoes" : "8 A 12",
                "descanso" : "1 MIN",
                "linkVideo" : "https://www.youtube.com/watch?v=p0Mql-QHnvY",
                "observacao" : "DROP SET NA ULTIMA SÉRIE"
            }),
            Exercicio({
                "nomeExercicio" : "EXTENSORA",
                "series" : "3",
                "repeticoes" : "8 A 12",
                "descanso" : "1 MIN",
                "linkVideo" : "https://www.youtube.com/watch?v=p0Mql-QHnvY",
                "observacao" : "DROP SET NA ULTIMA SÉRIE"
            }),
            Exercicio({
                "nomeExercicio" : "GLUTEO POLIA BAIXA",
                "series" : "4",
                "repeticoes" : "8 A 12",
                "descanso" : "1 MIN",
                "linkVideo" : "https://www.youtube.com/watch?v=p0Mql-QHnvY",
                "observacao" : "DROP SET NA ULTIMA SÉRIE"
            }),
            Exercicio({
                "nomeExercicio" : "PANTURRILHA SENTADA",
                "series" : "4",
                "repeticoes" : "8 A 12",
                "descanso" : "1 MIN",
                "linkVideo" : "https://www.youtube.com/watch?v=p0Mql-QHnvY",
                "observacao" : "DROP SET NA ULTIMA SÉRIE"
            }),
        ]
    }

    exerciciosQuinta = {
        "dia" : "QUI",
        "exercicios" :[
            Exercicio({
                "nomeExercicio" : "VOADOR",
                "series" : "4",
                "repeticoes" : "8 A 12",
                "descanso" : "1 MIN",
                "linkVideo" : "https://www.youtube.com/watch?v=p0Mql-QHnvY",
                "observacao" : "DROP SET NA ULTIMA SÉRIE"
            }),
            Exercicio({
                "nomeExercicio" : "ELEVAÇÃO LATERAL",
                "series" : "4",
                "repeticoes" : "8 A 12",
                "descanso" : "1 MIN",
                "linkVideo" : "https://www.youtube.com/watch?v=p0Mql-QHnvY",
                "observacao" : "DROP SET NA ULTIMA SÉRIE"
            }),
            Exercicio({
                "nomeExercicio" : "LEG MAQUINA",
                "series" : "3",
                "repeticoes" : "8 A 12",
                "descanso" : "1 MIN",
                "linkVideo" : "https://www.youtube.com/watch?v=p0Mql-QHnvY",
                "observacao" : "DROP SET NA ULTIMA SÉRIE"
            }),
            Exercicio({
                "nomeExercicio" : "EXTENSORA",
                "series" : "3",
                "repeticoes" : "8 A 12",
                "descanso" : "1 MIN",
                "linkVideo" : "https://www.youtube.com/watch?v=p0Mql-QHnvY",
                "observacao" : "DROP SET NA ULTIMA SÉRIE"
            }),
            Exercicio({
                "nomeExercicio" : "GLUTEO POLIA BAIXA",
                "series" : "4",
                "repeticoes" : "8 A 12",
                "descanso" : "1 MIN",
                "linkVideo" : "https://www.youtube.com/watch?v=p0Mql-QHnvY",
                "observacao" : "DROP SET NA ULTIMA SÉRIE"
            }),
            Exercicio({
                "nomeExercicio" : "PANTURRILHA SENTADA",
                "series" : "4",
                "repeticoes" : "8 A 12",
                "descanso" : "1 MIN",
                "linkVideo" : "https://www.youtube.com/watch?v=p0Mql-QHnvY",
                "observacao" : "DROP SET NA ULTIMA SÉRIE"
            }),
        ]
    }

    exerciciosSexta = {
        "dia" : "SEX",
        "exercicios" :[
            Exercicio({
                "nomeExercicio" : "VOADOR",
                "series" : "4",
                "repeticoes" : "8 A 12",
                "descanso" : "1 MIN",
                "linkVideo" : "https://www.youtube.com/watch?v=p0Mql-QHnvY",
                "observacao" : "DROP SET NA ULTIMA SÉRIE"
            }),
            Exercicio({
                "nomeExercicio" : "ELEVAÇÃO LATERAL",
                "series" : "4",
                "repeticoes" : "8 A 12",
                "descanso" : "1 MIN",
                "linkVideo" : "https://www.youtube.com/watch?v=p0Mql-QHnvY",
                "observacao" : "DROP SET NA ULTIMA SÉRIE"
            }),
            Exercicio({
                "nomeExercicio" : "LEG MAQUINA",
                "series" : "3",
                "repeticoes" : "8 A 12",
                "descanso" : "1 MIN",
                "linkVideo" : "https://www.youtube.com/watch?v=p0Mql-QHnvY",
                "observacao" : "DROP SET NA ULTIMA SÉRIE"
            }),
            Exercicio({
                "nomeExercicio" : "EXTENSORA",
                "series" : "3",
                "repeticoes" : "8 A 12",
                "descanso" : "1 MIN",
                "linkVideo" : "https://www.youtube.com/watch?v=p0Mql-QHnvY",
                "observacao" : "DROP SET NA ULTIMA SÉRIE"
            }),
            Exercicio({
                "nomeExercicio" : "GLUTEO POLIA BAIXA",
                "series" : "4",
                "repeticoes" : "8 A 12",
                "descanso" : "1 MIN",
                "linkVideo" : "https://www.youtube.com/watch?v=p0Mql-QHnvY",
                "observacao" : "DROP SET NA ULTIMA SÉRIE"
            }),
            Exercicio({
                "nomeExercicio" : "PANTURRILHA SENTADA",
                "series" : "4",
                "repeticoes" : "8 A 12",
                "descanso" : "1 MIN",
                "linkVideo" : "https://www.youtube.com/watch?v=p0Mql-QHnvY",
                "observacao" : "DROP SET NA ULTIMA SÉRIE"
            }),
        ]
    }
    
    treinoA = Treino(exerciciosSegunda)
    treinoB = Treino(exerciciosTerca)
    treinoC = Treino(exerciciosQuarta)
    treinoD = Treino(exerciciosQuinta)
    treinoE = Treino(exerciciosSexta)
    treinos = [treinoA,treinoB,treinoC,treinoD,treinoE]
    posTreino = [
        {"nome":"30 MIN ESTEIRA ( VELOCIDADE MINIMA 6.0 )", "link":"https://www.youtube.com/watch?v=fR_KzGyKW3w&t=25s"},
        {"nome":"ABDOMINAL BICICLETA 3X40SEG", "link":"https://www.youtube.com/watch?v=fR_KzGyKW3w&t=25s"},
        {"nome":"5MIN CORRENDO ESTEIRA", "link":"https://www.youtube.com/watch?v=fR_KzGyKW3w&t=25s"},
        {"nome":"ABDOMINAL CRUNCH 3X25", "link":"https://www.youtube.com/watch?v=fR_KzGyKW3w&t=25s"},
        {"nome":"ELEVAÇÃO DE PERNAS 3X10", "link":"https://www.youtube.com/watch?v=fR_KzGyKW3w&t=25s"},
        {"nome":"PRANCHA ISOMETRICA 3XRM", "link":"https://www.youtube.com/watch?v=fR_KzGyKW3w&t=25s"},
        {"nome":"GIRO RUSSO 3X50SEG", "link":"https://www.youtube.com/watch?v=fR_KzGyKW3w&t=25s"},
    ]

    args = {
        "workbookName" : "Planejamento.xslx",
        "nomeAluno" : "Carlos",
        "objetivo" : "Ganho de massa muscular",
        "diasDaSemana" : " Indefinido ",
        "observacao" : "2º treino adaptativo",
        "dataDeInicio" : "10/12/2022",
        "treinos" : treinos,
        "tipoTreino" : 'Força',
        "posTreino": posTreino
    }

    sheet = WorksheetMaycao(args) 
    sheet.createSheet()
    sheet.setDimensions()
    sheet.appendDadosAluno()
    sheet.stylesDadosAluno()
    sheet.appendExercicios()
    sheet.appendPosTreino()
    sheet.stylePosTreino()
    sheet.workbook.save(filename = "Carlos trieno.xlsx")
    




if __name__ == "__main__":
    main()
    print('\n\t\t',datetime.today().strftime('%A, %B %d, %Y %H:%M:%S'))