from datetime import datetime
import openpyxl
from textwrap import fill
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Color, colors, Side, GradientFill, Alignment, Border, PatternFill

class WorksheetMaycao:
    def __init__(self, args):
        self.nomeAluno    = args["nomeAluno"]
        self.workbookName = args["workbookName"]
        self.objetivo     = args["objetivo"]
        self.diasDaSemana = args["diasDaSemana"]
        self.observacao   = args["observacao"]
        self.dataDeInicio = args["dataDeInicio"]
        self.treinos      = args['treinos']
        self.tipoTreino   = args['tipoTreino']
        self.posTreino    = args['posTreino']
        self.workbook     = 0
        self.worksheet    = 0
        self.lastLineOfEexercicio = 0

    def createSheet(self):
        self.workbook  = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = self.workbookName   

    def setDimensions(self):
        self.worksheet.column_dimensions['A'].width = 5 
        self.worksheet.column_dimensions['B'].width = 35 
        self.worksheet.column_dimensions['C'].width = 15 
        self.worksheet.column_dimensions['D'].width = 25 
        self.worksheet.column_dimensions['E'].width = 10 
        self.worksheet.column_dimensions['F'].width = 15 
        self.worksheet.column_dimensions['G'].width = 7 
        self.worksheet.column_dimensions['H'].width = 30
        self.worksheet.column_dimensions['I'].width = 10  
        self.worksheet.column_dimensions['A'].height = 25
        self.worksheet.column_dimensions['B'].height = 25
        self.worksheet.column_dimensions['C'].height = 25
        self.worksheet.column_dimensions['D'].height = 25
        self.worksheet.column_dimensions['E'].height = 25
        self.worksheet.column_dimensions['F'].height = 25
        self.worksheet.column_dimensions['G'].height = 25

    def appendDadosAluno(self): 
        self.worksheet['C1'] = "Aluno:"
        self.worksheet['C2'] = "Data de Ínicio:"
        self.worksheet['C3'] = "Objetivo:"
        self.worksheet['C4'] = "Dias da semana:"
        self.worksheet['C5'] = "Observação:"
        self.worksheet['C6'] = "Treino:"
        self.worksheet['D1'] = self.nomeAluno
        self.worksheet['D2'] = self.dataDeInicio
        self.worksheet['D3'] = self.objetivo
        self.worksheet['D4'] = self.diasDaSemana
        self.worksheet['D5'] = self.observacao
        self.worksheet['D6'] = self.tipoTreino
    
    def stylesDadosAluno(self):
        key_font = Font(color='00FFFFFF')
        key_fill = PatternFill("solid", fgColor="00969696")
        key_alignment = Alignment(horizontal="right", vertical="center")     
        self.worksheet['C1'].border = Border(top = Side(border_style="thick", color="00000000"),
        left = Side(border_style="thick", color="00000000"))
        self.worksheet['C2'].border = Border(left = Side(border_style="thick", color="00000000"))
        self.worksheet['C3'].border = Border(left = Side(border_style="thick", color="00000000"))
        self.worksheet['C4'].border = Border(left = Side(border_style="thick", color="00000000"))
        self.worksheet['C5'].border = Border(left = Side(border_style="thick", color="00000000"))
        self.worksheet['D2'].border = Border(right = Side(border_style="thick", color="00000000"))
        self.worksheet['D3'].border = Border(right = Side(border_style="thick", color="00000000"))
        self.worksheet['D4'].border = Border(right = Side(border_style="thick", color="00000000"))
        self.worksheet['D5'].border = Border(right = Side(border_style="thick", color="00000000"))
        self.worksheet['C6'].border = Border(left = Side(border_style="thick", color="00000000"),
                                            bottom = Side(border_style="thick", color="00000000"))
        self.worksheet['D1'].border = Border(top = Side(border_style="thick", color="00000000"),
                                            right = Side(border_style="thick", color="00000000"))
        self.worksheet['D6'].border = Border(right = Side(border_style="thick", color="00000000"),
                                            bottom = Side(border_style="thick", color="00000000"))
        count = 1
        while count < 7:
            cel = 'C' + str(count)
            self.worksheet[cel].font = key_font
            self.worksheet[cel].fill = key_fill
            self.worksheet[cel].alignment = key_alignment
            count = count + 1
        value_font = Font(color='00000000', bold=True)
        value_fill = PatternFill("solid", fgColor="00FFFFFF")
        value_alignment = Alignment(horizontal="center", vertical="center")
        count = 1
        while count < 7:
            cel = 'D' + str(count)
            self.worksheet[cel].font = value_font
            self.worksheet[cel].fill = value_fill
            self.worksheet[cel].alignment = value_alignment
            count = count + 1

    def appendExercicios(self):
        startingCelLine = 7
        columnNumber = str(startingCelLine)
        for i, treino in enumerate(self.treinos):
            startingCelLine+= 1
            treinoStartingCelLineCopy = startingCelLine + 1
            columnNumber = str(startingCelLine)
            self.worksheet['A' + columnNumber] = "Nº"
            self.worksheet['B' + columnNumber] = "EXERCÍCIO"
            self.worksheet['C' + columnNumber] = "SÉRIES"
            self.worksheet['D' + columnNumber] = "REPETIÇÕES"
            self.worksheet['E' + columnNumber] = "CARGA KG"
            self.worksheet['F' + columnNumber] = "DESCANSO"
            self.worksheet['G' + columnNumber] = "VÍDEO"
            self.worksheet['H' + columnNumber] = "OBS"
            self.worksheet['I' + columnNumber] = "VOLUME"
            self.styleExercicioColumns(columnNumber)
            for iExercicio, exercicio in enumerate(treino.exercicios):
                startingCelLine+= 1
                columnNumber = str(startingCelLine)
                self.worksheet['A' + columnNumber] = iExercicio + 1
                self.worksheet['B' + columnNumber] = exercicio.nomeExercicio
                self.worksheet['C' + columnNumber] = exercicio.series
                self.worksheet['D' + columnNumber] = exercicio.repeticoes
                self.worksheet['E' + columnNumber] = exercicio.carga
                self.worksheet['F' + columnNumber] = exercicio.descanso
                self.worksheet['G' + columnNumber].hyperlink = exercicio.linkVideo
                self.worksheet['G' + columnNumber].value = ''
                img = openpyxl.drawing.image.Image('./img/ytb.png')
                img.width = 20
                img.height = 20
                img.anchor = 'G' + columnNumber
                self.worksheet.add_image(img)
                self.worksheet['H' + columnNumber] = exercicio.observacao
                self.styleExercicioRows(columnNumber)
                treinoEndingCelLineCopy = startingCelLine -1
                warningErrorFormula = '\"Dados inválidos\"'
                formulaSimples = f"=IFERROR(D{columnNumber}*C{columnNumber}*E{columnNumber},{warningErrorFormula})"
                self.worksheet['I' + columnNumber].value = formulaSimples
            self.worksheet.merge_cells('J' + str(treinoStartingCelLineCopy) + ':J' + str(columnNumber))
            self.worksheet['J' + str(treinoStartingCelLineCopy)].alignment = Alignment(horizontal="center", vertical="center")
            self.worksheet['J' + str(treinoStartingCelLineCopy)].fill = PatternFill("solid", fgColor="00333333")
            self.worksheet['J' + str(treinoStartingCelLineCopy)].font = Font(color='00FFFF00',size=20)
            self.worksheet['J' + str(treinoStartingCelLineCopy)] = treino.dia   
        self.lastLineOfEexercicio = int(columnNumber)     

    def styleExercicioRows(self, line) : 
        alignment = Alignment(horizontal="center", vertical="center")
        coluns = ['A','B','C','D','E','F','G', 'H','I','J']
        for column in coluns : 
            self.worksheet[column + line].alignment = alignment

    def styleExercicioColumns(self, line) :
        fill = PatternFill("solid", fgColor="00333333")
        font = Font(color='00FFFF00')
        alignment = Alignment(horizontal="center", vertical="center")
        coluns = ['A','B','C','D','E','F','G', 'H','I','J']
        
        for column in coluns:
            if(column != 'J'):
                self.worksheet[column + line].fill = fill
            self.worksheet[column + line].font = font
            self.worksheet[column + line].alignment = alignment
        
          
    def appendPosTreino(self):
        startingLine = self.lastLineOfEexercicio + 3;
        count = startingLine;
        endLine = startingLine + len(self.posTreino)
        titleCel = 'A' + str( startingLine - 1 )  
        self.worksheet[titleCel] = 'PÓS TREINO TODOS OS DIAS'
        while count != endLine :
            cellCount = 'A' + str(count)
            cellNome  = 'B' + str(count) 
            cellLink  = 'C' + str(count) 
            exercicio = self.posTreino[count - startingLine]
            self.worksheet[cellCount] = count - startingLine + 1
            self.worksheet[cellNome] = exercicio['descricao']
            print('aqui')
            print(exercicio['link'])
            self.worksheet[cellLink].value = f"=HYPERLINK(\"{exercicio['link']}\",\"\")"
            img = openpyxl.drawing.image.Image('./img/ytb.png')
            img.width = 20
            img.height = 20
            img.anchor = cellLink
            self.worksheet.add_image(img)
            count = count + 1
        return 1;

    def stylePosTreino(self):
        startingLine = self.lastLineOfEexercicio + 2;
        count = startingLine
        endLine = startingLine + 1 + len(self.posTreino)
        fill = PatternFill("solid", fgColor="00333333")
        font = Font(color='00FFFF00')
        alignment = Alignment(horizontal="center", vertical="center")
        coluns = ['A','B','C']
        for colunm in coluns : 
            self.worksheet[colunm + str(startingLine)].fill = fill
            self.worksheet[colunm + str(startingLine)].font = font
            self.worksheet[colunm + str(startingLine)].alignment = alignment

        while count != endLine:
            celA = "A" + str(count)
            celB = 'B' + str(count)
            celC = 'D' + str(count)
            self.worksheet[celA].alignment = alignment
            self.worksheet[celB].alignment = alignment
            self.worksheet[celC].alignment = alignment
            count = count + 1
        self.worksheet.merge_cells('A' + str(startingLine) + ":C" + str(startingLine))
       

